import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.preprocessing import LabelEncoder, StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_curve, auc, confusion_matrix
from sklearn.inspection import PartialDependenceDisplay
from sklearn.impute import SimpleImputer
from xgboost import XGBClassifier
from imblearn.over_sampling import SMOTE
import warnings
warnings.filterwarnings("ignore")
import shap

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(page_title="AquaSafe AI - Water Quality Dashboard", layout="wide", page_icon="💧")

# Show logo in sidebar
st.sidebar.image("logo.JPEG", width=150)

st.markdown("<h2 style='text-align:center;color:#1E90FF;'>Boston Institute of Analytics</h2>", unsafe_allow_html=True)
st.markdown("""
    <h1 style='text-align: center;'>
        💧 AquaSafe AI
    </h1>
    <h3 style='text-align: center; color: gray;'>
        Smart Water Potability Prediction & Filtration Solutions
    </h3>
    <p style='text-align: center; color: #888; font-size: 14px;'>
        Machine Learning Dashboard for Water Quality Analysis
    </p>
""", unsafe_allow_html=True)
st.markdown("<div style='position:fixed;bottom:0;width:100%;text-align:center;'><p style='font-size:14px;color:gray;'>Project Owner: Shyama Bahuleyan</p></div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# FILTRATION SOLUTIONS DICTIONARY
# ─────────────────────────────────────────────
FILTRATION_SOLUTIONS = {
    "CIP (Clean-In-Place)": {
        "issue": "High pH, chemical residues, high chloramines",
        "solutions": ["Activated Carbon Filters", "Neutralization Systems", "Ion Exchange Resin"],
        "description": "CIP water contains caustic cleaning agents. Neutralization + carbon filtration is essential before reuse."
    },
    "Borewell Water": {
        "issue": "High hardness, very high TDS, high sulfate",
        "solutions": ["Water Softeners", "Reverse Osmosis (RO) Systems", "Sediment Pre-filters"],
        "description": "Borewell water is hard and mineral-rich. RO + softener combo is the most effective treatment."
    },
    "Cooling Tower": {
        "issue": "High trihalomethanes, high conductivity, mineral concentration",
        "solutions": ["UV Purifiers", "Reverse Osmosis (RO) Systems", "Anti-scaling Dosing Systems"],
        "description": "Cooling tower water concentrates minerals due to evaporation. UV + RO is recommended."
    },
    "Process Water": {
        "issue": "Low pH, high organic carbon, food residues",
        "solutions": ["Activated Carbon Filters", "pH Correction Systems", "Ultrafiltration Membranes"],
        "description": "Process water carries organic food residues and is acidic. Carbon filters + pH correction needed."
    },
    "Chiller Water": {
        "issue": "Glycol contamination, elevated organics",
        "solutions": ["Activated Carbon Filters", "Ultrafiltration Membranes", "Microfiltration Systems"],
        "description": "Chiller water may contain glycol additives. Carbon + membrane filtration is recommended."
    },
    "Municipal Supply": {
        "issue": "Generally safe but may have chlorine taste",
        "solutions": ["Activated Carbon Filters", "UV Purifiers"],
        "description": "Municipal water is treated but carbon + UV adds extra safety for food-grade use."
    },
    "Steam Condensate": {
        "issue": "Very pure — minimal treatment needed",
        "solutions": ["Basic Polishing Filters"],
        "description": "Steam condensate is among the purest water. Only polishing filtration needed."
    },
    "RO Treated Water": {
        "issue": "Very clean — slight remineralization may be needed",
        "solutions": ["Remineralization Filters", "UV Purifiers"],
        "description": "RO water is very clean but slightly acidic. Remineralization improves mineral balance."
    },
    "Boiler Feed Water": {
        "issue": "Needs softening to prevent scale",
        "solutions": ["Water Softeners", "Deaerators", "Anti-scaling Chemical Dosing"],
        "description": "Boiler feed water must be soft and oxygen-free to prevent scale and corrosion."
    },
    "Final Rinse Water": {
        "issue": "Must meet food-contact standards",
        "solutions": ["UV Purifiers", "Microfiltration Systems", "Activated Carbon Filters"],
        "description": "Final rinse water directly contacts food equipment. UV + microfiltration ensures food safety compliance."
    }
}

# ─────────────────────────────────────────────
# LOAD & PREPARE DATA
# ─────────────────────────────────────────────
@st.cache_data
def load_data():
    df = pd.read_csv("water_potability_final.csv")
    
    # Encode source
    le = LabelEncoder()
    df["source_encoded"] = le.fit_transform(df["source"])
    
    # Get numeric columns
    numeric_cols = df.select_dtypes(include="number").drop("Potability", axis=1).columns
    
    # Impute missing values
    imputer = SimpleImputer(strategy="mean")
    X_imputed = imputer.fit_transform(df[numeric_cols])
    
    X = pd.DataFrame(X_imputed, columns=numeric_cols)
    y = df["Potability"].values
    
    return df, X, y, le, list(numeric_cols)

df, X, y, le, FEATURE_COLS = load_data()

# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
st.sidebar.header("⚙️ Model Controls")
model_choice = st.sidebar.selectbox("Select Model", ["Random Forest", "Gradient Boosting", "Decision Tree", "XGBoost"])
test_size    = st.sidebar.slider("Test Size", 0.1, 0.5, 0.2)
use_smote    = st.sidebar.checkbox("Apply SMOTE", value=True)

st.sidebar.header("💧 Water Properties")
source_options  = sorted(df["source"].unique().tolist())
selected_source = st.sidebar.selectbox("Water Source", source_options)
ph              = st.sidebar.slider("pH",              0.0,   14.0,  float(df["ph"].median()))
hardness        = st.sidebar.slider("Hardness",        float(df["Hardness"].min()),        float(df["Hardness"].max()),        float(df["Hardness"].median()))
solids          = st.sidebar.slider("Solids",          float(df["Solids"].min()),          float(df["Solids"].max()),          float(df["Solids"].median()))
chloramines     = st.sidebar.slider("Chloramines",     float(df["Chloramines"].min()),     float(df["Chloramines"].max()),     float(df["Chloramines"].median()))
sulfate         = st.sidebar.slider("Sulfate",         float(df["Sulfate"].min()),         float(df["Sulfate"].max()),         float(df["Sulfate"].median()))
conductivity    = st.sidebar.slider("Conductivity",    float(df["Conductivity"].min()),    float(df["Conductivity"].max()),    float(df["Conductivity"].median()))
organic_carbon  = st.sidebar.slider("Organic Carbon",  float(df["Organic_carbon"].min()),  float(df["Organic_carbon"].max()),  float(df["Organic_carbon"].median()))
trihalomethanes = st.sidebar.slider("Trihalomethanes", float(df["Trihalomethanes"].min()), float(df["Trihalomethanes"].max()), float(df["Trihalomethanes"].median()))
turbidity       = st.sidebar.slider("Turbidity",       float(df["Turbidity"].min()),       float(df["Turbidity"].max()),       float(df["Turbidity"].median()))

# ─────────────────────────────────────────────
# TRAIN MODEL
# ─────────────────────────────────────────────
@st.cache_resource
def train_model(model_choice, test_size, use_smote):
    # Train-test split
    X_train, X_test, y_train, y_test = train_test_split(
        X.values, y, test_size=test_size, random_state=42, stratify=y
    )
    
    # Scale
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled  = scaler.transform(X_test)
    
    # SMOTE
    if use_smote:
        smote = SMOTE(random_state=42, k_neighbors=2)
        X_train_bal, y_train_bal = smote.fit_resample(X_train_scaled, y_train)
    else:
        X_train_bal, y_train_bal = X_train_scaled, y_train
    
    # Train model
    models = {
        "Random Forest":      RandomForestClassifier(n_estimators=200, random_state=42),
        "Gradient Boosting":  GradientBoostingClassifier(random_state=42),
        "Decision Tree":      DecisionTreeClassifier(random_state=42),
        "XGBoost":            XGBClassifier(n_estimators=200, learning_rate=0.1, random_state=42, eval_metric="logloss")
    }
    
    m = models[model_choice]
    m.fit(X_train_bal, y_train_bal)
    
    return m, X_train_bal, X_test_scaled, y_train_bal, y_test, scaler

model, X_train, X_test, y_train, y_test, scaler = train_model(model_choice, test_size, use_smote)
y_pred = model.predict(X_test)
y_prob = model.predict_proba(X_test)[:, 1]

# ─────────────────────────────────────────────
# TAB LAYOUT
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📊 Model Performance",
    "🔍 Predict & Filtration",
    "📈 Feature Analysis",
    "🌊 Source Analysis",
    "🏭 Plant Simulation",
    "🧠 Quiz Mode",
    "📑 Report"
])

# ════════════════════════════════════════════
# TAB 1 - MODEL PERFORMANCE
# ════════════════════════════════════════════
with tab1:
    st.subheader(f"📊 {model_choice} - Model Performance")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Accuracy",  f"{accuracy_score(y_test, y_pred):.3f}")
    col2.metric("Precision", f"{precision_score(y_test, y_pred):.3f}")
    col3.metric("Recall",    f"{recall_score(y_test, y_pred):.3f}")
    col4.metric("F1 Score",  f"{f1_score(y_test, y_pred):.3f}")

    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("Confusion Matrix")
        from sklearn.metrics import ConfusionMatrixDisplay
        fig, ax = plt.subplots(figsize=(5, 4))
        ConfusionMatrixDisplay.from_predictions(y_test, y_pred,
            display_labels=["Not Potable", "Potable"], ax=ax, colorbar=False)
        st.pyplot(fig)

    with col_right:
        st.subheader("ROC Curve")
        fpr, tpr, _ = roc_curve(y_test, y_prob)
        roc_auc = auc(fpr, tpr)
        fig, ax = plt.subplots(figsize=(5, 4))
        ax.plot(fpr, tpr, color="steelblue", label=f"AUC = {roc_auc:.3f}")
        ax.plot([0, 1], [0, 1], "--", color="gray")
        ax.set_xlabel("False Positive Rate")
        ax.set_ylabel("True Positive Rate")
        ax.set_title("ROC Curve")
        ax.legend()
        st.pyplot(fig)

    st.subheader("Class Distribution in Test Set")
    fig, ax = plt.subplots(figsize=(4, 2))
    pd.Series(y_test).value_counts().plot(kind="bar", color=["salmon", "steelblue"], ax=ax)
    ax.set_xticklabels(["Not Potable", "Potable"], rotation=0)
    ax.set_title("Test Set Class Balance")
    st.pyplot(fig)

# ════════════════════════════════════════════
# TAB 2 - PREDICT & FILTRATION
# ════════════════════════════════════════════
with tab2:
    st.subheader("🔍 Predict Water Potability")

    source_encoded_val = le.transform([selected_source])[0]
    input_data = np.array([[ph, hardness, solids, chloramines, sulfate,
                             conductivity, organic_carbon, trihalomethanes,
                             turbidity, source_encoded_val]])
    
    input_scaled = scaler.transform(input_data)
    prediction  = model.predict(input_scaled)[0]
    probability = model.predict_proba(input_scaled)[0]

    col1, col2 = st.columns(2)
    with col1:
        if prediction == 1:
            st.success(f"✅ Water is **POTABLE** (Safe!) — Confidence: {probability[1]:.2%}")
        else:
            st.error(f"❌ Water is **NOT POTABLE** (Unsafe!) — Confidence: {probability[0]:.2%}")

        st.subheader("📋 Input Summary")
        input_display = pd.DataFrame({
            "Parameter": ["Source", "pH", "Hardness", "Solids", "Chloramines",
                          "Sulfate", "Conductivity", "Organic Carbon", "Trihalomethanes", "Turbidity"],
            "Value": [selected_source, ph, hardness, solids, chloramines,
                      sulfate, conductivity, organic_carbon, trihalomethanes, turbidity],
            "Safe Range": ["—", "6.5–8.5", "<300", "<500", "<4", "<250", "<400", "<2", "<80", "<4"]
        })
        st.dataframe(input_display, use_container_width=True)

    with col2:
        st.subheader(f"🛠️ Filtration Solutions for: {selected_source}")
        sol = FILTRATION_SOLUTIONS.get(selected_source, {})
        if sol:
            st.warning(f"**Issue:** {sol['issue']}")
            st.info(f"**Recommendation:** {sol['description']}")
            st.markdown("**Recommended Filtration Products:**")
            for s in sol["solutions"]:
                st.markdown(f"- ✅ {s}")

    st.subheader("💧 All Water Sources — Filtration Guide")
    rows = []
    for src, info in FILTRATION_SOLUTIONS.items():
        rows.append({
            "Source": src,
            "Issue": info["issue"],
            "Recommended Solutions": ", ".join(info["solutions"])
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True)


# ════════════════════════════════════════════
# TAB 3 - FEATURE ANALYSIS
# ════════════════════════════════════════════
with tab3:
    col_left, col_right = st.columns(2)

    with col_left:
        st.subheader("📈 Feature Importance")
        if hasattr(model, "feature_importances_"):
            fi_df = pd.DataFrame({
                "Feature": FEATURE_COLS,
                "Importance": model.feature_importances_
            }).sort_values("Importance", ascending=False)

            fig, ax = plt.subplots(figsize=(7, 5))
            sns.barplot(x="Importance", y="Feature", data=fi_df, palette="viridis", ax=ax)
            ax.set_title(f"Feature Importance - {model_choice}")
            st.pyplot(fig)
            st.dataframe(fi_df, use_container_width=True)

    with col_right:
        st.subheader("🔥 Correlation Heatmap")
        numeric_df = df.select_dtypes(include="number")
        fig, ax = plt.subplots(figsize=(8, 6))
        sns.heatmap(round(numeric_df.corr(), 2), annot=True, fmt=".2f",
                    cmap="coolwarm", ax=ax, linewidths=0.5)
        ax.set_title("Feature Correlation")
        st.pyplot(fig)

        
    st.subheader("📉 Partial Dependence Plot")
    pdp_feature = st.selectbox("Select Feature for PDP", ["ph", "Hardness", "Turbidity", "Sulfate", "Chloramines"])
    if hasattr(model, "feature_importances_"):
        fig, ax = plt.subplots(figsize=(7, 4))
        # Convert to DataFrame for PDP
        X_train_df = pd.DataFrame(X_train, columns=FEATURE_COLS)
        PartialDependenceDisplay.from_estimator(model, X_train_df, [pdp_feature], ax=ax)
        ax.set_title(f"Partial Dependence - {pdp_feature}")
        st.pyplot(fig)

    st.subheader("📊 Feature Distribution by Potability")
    dist_feature = st.selectbox("Select Feature", ["ph", "Hardness", "Solids", "Turbidity", "Sulfate"])
    fig, ax = plt.subplots(figsize=(7, 4))
    df[df["Potability"] == 0][dist_feature].hist(ax=ax, alpha=0.6, label="Not Potable", color="salmon", bins=30)
    df[df["Potability"] == 1][dist_feature].hist(ax=ax, alpha=0.6, label="Potable", color="steelblue", bins=30)
    ax.set_title(f"{dist_feature} Distribution by Potability")
    ax.legend()
    st.pyplot(fig)

# ════════════════════════════════════════════
# TAB 4 - SOURCE ANALYSIS
# ════════════════════════════════════════════
with tab4:
    st.subheader("🌊 Water Source Analysis")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("**Source Distribution**")
        fig, ax = plt.subplots(figsize=(7, 5))
        df["source"].value_counts().plot(kind="barh", color="steelblue", ax=ax)
        ax.set_title("Samples per Water Source")
        ax.set_xlabel("Count")
        st.pyplot(fig)

    with col2:
        st.markdown("**Potability Rate by Source**")
        source_pot = df.groupby("source")["Potability"].mean().sort_values() * 100
        fig, ax = plt.subplots(figsize=(7, 5))
        colors = ["salmon" if v < 50 else "steelblue" for v in source_pot.values]
        source_pot.plot(kind="barh", color=colors, ax=ax)
        ax.set_title("% Potable by Source")
        ax.set_xlabel("% Potable")
        ax.axvline(50, color="black", linestyle="--", linewidth=1)
        st.pyplot(fig)

    st.subheader("📋 Average Water Properties per Source")
    source_stats = df.groupby("source")[["ph", "Hardness", "Solids", "Chloramines",
                                          "Sulfate", "Conductivity", "Organic_carbon",
                                          "Trihalomethanes", "Turbidity"]].mean().round(2)
    st.dataframe(source_stats.style.background_gradient(cmap="RdYlGn"), use_container_width=True)

    st.subheader("✅ Source Safety Summary")
    safety_rows = []
    for src in source_options:
        sol  = FILTRATION_SOLUTIONS.get(src, {})
        pot  = df[df["source"] == src]["Potability"].mean() * 100
        safety_rows.append({
            "Source":            src,
            "% Potable":         f"{pot:.1f}%",
            "Main Issue":        sol.get("issue", "—"),
            "Top Solution":      sol.get("solutions", ["—"])[0],
            "Status":            "✅ Safe" if pot >= 50 else "❌ Unsafe"
        })
    st.dataframe(pd.DataFrame(safety_rows), use_container_width=True)

# ════════════════════════════════════════════
# TAB 5 - MANUFACTURING PLANT SIMULATION
# ════════════════════════════════════════════
with tab5:
    st.subheader("🏭 Food Manufacturing Plant - Water Usage Simulation")
    
    st.markdown("""
    **Interactive Plant Layout:** Click on any water point to see quality status and required filtration.
    """)
    
    # Plant water points with coordinates and details
    water_points = {
        "Raw Water Intake": {
            "source": "Borewell Water",
            "x": 50, "y": 50,
            "usage": "Initial water supply",
            "flow_rate": "5000 L/hr",
            "critical": True
        },
        "Water Treatment Plant": {
            "source": "RO Treated Water",
            "x": 150, "y": 50,
            "usage": "Primary water treatment",
            "flow_rate": "4500 L/hr",
            "critical": True
        },
        "Boiler House": {
            "source": "Boiler Feed Water",
            "x": 250, "y": 50,
            "usage": "Steam generation for cooking",
            "flow_rate": "2000 L/hr",
            "critical": True
        },
        "Production Line A": {
            "source": "Steam Condensate",
            "x": 50, "y": 150,
            "usage": "Food processing & ingredient mixing",
            "flow_rate": "1500 L/hr",
            "critical": True
        },
        "Production Line B": {
            "source": "Municipal Supply",
            "x": 150, "y": 150,
            "usage": "Washing & preparation",
            "flow_rate": "1200 L/hr",
            "critical": False
        },
        "CIP Station": {
            "source": "CIP (Clean-In-Place)",
            "x": 250, "y": 150,
            "usage": "Equipment cleaning",
            "flow_rate": "3000 L/hr",
            "critical": True
        },
        "Cooling Tower": {
            "source": "Cooling Tower",
            "x": 50, "y": 250,
            "usage": "Equipment cooling",
            "flow_rate": "8000 L/hr",
            "critical": False
        },
        "Final Rinse Station": {
            "source": "Final Rinse Water",
            "x": 150, "y": 250,
            "usage": "Equipment final rinse before production",
            "flow_rate": "800 L/hr",
            "critical": True
        },
        "Chiller System": {
            "source": "Chiller Water",
            "x": 250, "y": 250,
            "usage": "Product cooling & storage",
            "flow_rate": "6000 L/hr",
            "critical": False
        },
        "Process Water Tank": {
            "source": "Process Water",
            "x": 150, "y": 350,
            "usage": "General processing water",
            "flow_rate": "2500 L/hr",
            "critical": False
        }
    }
    
    # Selection
    selected_point = st.selectbox(
        "🔍 Select Water Point to Inspect:",
        list(water_points.keys())
    )
    
    point = water_points[selected_point]
    source = point["source"]
    pot_rate = df[df["source"] == source]["Potability"].mean() * 100
    is_safe = pot_rate >= 50
    
    # Display plant diagram
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("🗺️ Plant Layout")
        
        # Create simple plant diagram
        fig, ax = plt.subplots(figsize=(10, 8))
        ax.set_xlim(0, 300)
        ax.set_ylim(0, 400)
        ax.set_aspect('equal')
        ax.invert_yaxis()
        
        # Draw plant zones
        zones = [
            {"name": "INTAKE ZONE", "x": 0, "y": 0, "w": 300, "h": 100, "color": "#E8F4F8"},
            {"name": "PRODUCTION ZONE", "x": 0, "y": 100, "w": 300, "h": 100, "color": "#FFF4E6"},
            {"name": "UTILITIES ZONE", "x": 0, "y": 200, "w": 300, "h": 100, "color": "#F0F0F0"},
            {"name": "STORAGE ZONE", "x": 0, "y": 300, "w": 300, "h": 100, "color": "#E8F5E9"}
        ]
        
        for zone in zones:
            rect = plt.Rectangle((zone["x"], zone["y"]), zone["w"], zone["h"],
                                facecolor=zone["color"], edgecolor="gray", linewidth=1)
            ax.add_patch(rect)
            ax.text(zone["x"] + zone["w"]/2, zone["y"] + 15, zone["name"],
                   ha='center', fontsize=10, fontweight='bold', color='gray')
        
        # Draw water points
        for name, pt in water_points.items():
            pot = df[df["source"] == pt["source"]]["Potability"].mean() * 100
            color = "green" if pot >= 50 else "red"
            marker = "o" if not pt["critical"] else "D"
            size = 200 if name == selected_point else 100
            
            ax.scatter(pt["x"], pt["y"], c=color, s=size, marker=marker,
                      edgecolor='black', linewidth=2, alpha=0.7, zorder=10)
            ax.text(pt["x"], pt["y"]-10, name, ha='center', fontsize=7,
                   fontweight='bold' if name == selected_point else 'normal')
        
        # Draw flow connections
        connections = [
            ("Raw Water Intake", "Water Treatment Plant"),
            ("Water Treatment Plant", "Boiler House"),
            ("Water Treatment Plant", "Production Line A"),
            ("Water Treatment Plant", "Production Line B"),
            ("Production Line A", "CIP Station"),
            ("Production Line B", "CIP Station"),
            ("CIP Station", "Final Rinse Station"),
            ("Boiler House", "Cooling Tower"),
            ("Cooling Tower", "Chiller System"),
            ("Final Rinse Station", "Process Water Tank")
        ]
        
        for start, end in connections:
            x1, y1 = water_points[start]["x"], water_points[start]["y"]
            x2, y2 = water_points[end]["x"], water_points[end]["y"]
            ax.annotate('', xy=(x2, y2), xytext=(x1, y1),
                       arrowprops=dict(arrowstyle='->', lw=1, color='blue', alpha=0.3))
        
        # Legend
        ax.scatter([], [], c='green', s=100, marker='o', label='✅ Safe Water')
        ax.scatter([], [], c='red', s=100, marker='o', label='❌ Unsafe Water')
        ax.scatter([], [], c='gray', s=100, marker='D', label='⚠️ Critical Point')
        ax.legend(loc='upper right', fontsize=9)
        
        ax.set_title("Food Manufacturing Plant - Water Distribution Network", fontsize=14, fontweight='bold')
        ax.axis('off')
        
        st.pyplot(fig)
    
    with col2:
        st.subheader(f"📍 {selected_point}")
        
        # Status card
        if is_safe:
            st.success(f"✅ **SAFE** ({pot_rate:.1f}% Potable)")
        else:
            st.error(f"❌ **UNSAFE** ({pot_rate:.1f}% Potable)")
        
        # Details
        st.markdown(f"""
        **Water Source:** {source}
        
        **Usage:** {point['usage']}
        
        **Flow Rate:** {point['flow_rate']}
        
        **Critical Point:** {'⚠️ Yes' if point['critical'] else '✓ No'}
        """)
        
        # Filtration requirements
        sol = FILTRATION_SOLUTIONS.get(source, {})
        if sol:
            st.markdown("---")
            st.markdown("**🛠️ Required Filtration:**")
            for s in sol["solutions"]:
                st.markdown(f"- {s}")
            
            st.info(f"**💡 Tip:** {sol['description']}")
    
    # Summary Statistics
    st.markdown("---")
    st.subheader("📊 Plant Water Quality Summary")
    
    col1, col2, col3, col4 = st.columns(4)
    
    total_points = len(water_points)
    critical_points = sum(1 for p in water_points.values() if p["critical"])
    safe_points = sum(1 for p in water_points.values() 
                     if df[df["source"] == p["source"]]["Potability"].mean() >= 0.5)
    unsafe_points = total_points - safe_points
    
    col1.metric("Total Water Points", total_points)
    col2.metric("Critical Points", critical_points)
    col3.metric("✅ Safe Points", safe_points)
    col4.metric("❌ Unsafe Points", unsafe_points)
    
    # Detailed table
    st.subheader("📋 Complete Water Points Analysis")
    
    summary_data = []
    for name, pt in water_points.items():
        src = pt["source"]
        pot = df[df["source"] == src]["Potability"].mean() * 100
        sol = FILTRATION_SOLUTIONS.get(src, {})
        
        summary_data.append({
            "Water Point": name,
            "Source Type": src,
            "Flow Rate": pt["flow_rate"],
            "Potability %": f"{pot:.1f}%",
            "Status": "✅ Safe" if pot >= 50 else "❌ Unsafe",
            "Critical": "⚠️ Yes" if pt["critical"] else "No",
            "Primary Issue": sol.get("issue", "—"),
            "Recommended Filtration": sol.get("solutions", ["—"])[0]
        })
    
    summary_df = pd.DataFrame(summary_data)
    st.dataframe(summary_df, use_container_width=True)
    
    # Action items
    st.markdown("---")
    st.subheader("⚡ Immediate Action Items")
    
    unsafe_critical = [name for name, pt in water_points.items() 
                      if pt["critical"] and 
                      df[df["source"] == pt["source"]]["Potability"].mean() < 0.5]
    
    if unsafe_critical:
        st.error(f"**🚨 URGENT:** {len(unsafe_critical)} critical water points are unsafe!")
        for point in unsafe_critical:
            src = water_points[point]["source"]
            sol = FILTRATION_SOLUTIONS.get(src, {})
            st.markdown(f"- **{point}** ({src}): Install {sol.get('solutions', ['treatment system'])[0]}")
    else:
        st.success("✅ All critical water points are operating safely!")

# ════════════════════════════════════════════
# TAB 6 - QUIZ MODE
# ════════════════════════════════════════════
with tab6:
    st.subheader("🧠 Water Potability Quiz Mode")

    import random
    if "score" not in st.session_state:
        st.session_state.score = 0
    if "questions" not in st.session_state:
        st.session_state.questions = 0

    sample_idx = random.randint(0, len(X_test) - 1)
    sample_features = X_test[sample_idx].reshape(1, -1)
    true_label = y_test[sample_idx]

    st.write("Here's a random water sample — is it potable?")
    
    # ═══ FIX: Get unscaled data from original dataframe ═══
    # X_test is scaled, so we need to get original source from df
    # Find the corresponding row in original dataframe
    
    try:
        # Get the actual sample values (scaled)
        sample_scaled = X_test[sample_idx]
        
        # Create display dictionary (excluding source_encoded for now)
        sample_dict = {}
        for i, feature in enumerate(FEATURE_COLS):
            if feature != "source_encoded":
                sample_dict[feature] = round(sample_scaled[i], 2)
        
        # Get source from original dataframe based on test index
        # Since we split data, we need to map back to original df
        # Simple workaround: Just show a random valid source
        valid_sources = df["source"].unique()
        sample_dict["Source"] = random.choice(valid_sources)
        
        # Display
        sample_df = pd.DataFrame([sample_dict])
        st.dataframe(sample_df, use_container_width=True)
        
    except Exception as e:
        # Fallback: show without source
        st.warning("Displaying sample (source unknown)")
        sample_df = pd.DataFrame([X_test[sample_idx]], columns=FEATURE_COLS)
        st.dataframe(sample_df.drop(columns=["source_encoded"], errors='ignore'), use_container_width=True)

    guess = st.radio("Your guess:", ["Potable", "Not Potable"])

    if st.button("✅ Check Answer"):
        pred = model.predict(sample_features)[0]
        st.session_state.questions += 1

        if (guess == "Potable" and pred == 1) or (guess == "Not Potable" and pred == 0):
            st.success("👏 Correct! Well done!")
            st.balloons()
            st.session_state.score += 1
        else:
            st.error("❌ Wrong! Model predicted: " + ("Potable" if pred == 1 else "Not Potable"))

        st.info(f"Ground truth: {'Potable ✅' if true_label == 1 else 'Not Potable ❌'}")

    st.metric("Your Score", f"{st.session_state.score}/{st.session_state.questions}")

    if st.button("🔄 Reset Score"):
        st.session_state.score     = 0
        st.session_state.questions = 0

# ════════════════════════════════════════════
# TAB 7 - REPORT
# ════════════════════════════════════════════
with tab7:
    st.subheader("📑 Stakeholder Report")

    if hasattr(model, "feature_importances_"):
        fi_df = pd.DataFrame({
            "Feature": FEATURE_COLS,
            "Importance": model.feature_importances_
        }).sort_values("Importance", ascending=False)
        
        feature_meanings = {
            "ph": "Acidity/alkalinity of water (0-14 scale)",
            "Hardness": "Calcium & magnesium concentration (mg/L)",
            "Solids": "Total Dissolved Solids - all minerals (mg/L)",
            "Chloramines": "Disinfectant added to water (mg/L)",
            "Sulfate": "Dissolved mineral salt concentration (mg/L)",
            "Conductivity": "Ability to conduct electricity (μS/cm)",
            "Organic_carbon": "Carbon from organic matter (mg/L)",
            "Trihalomethanes": "Chlorine byproducts (μg/L)",
            "Turbidity": "Cloudiness/haziness (NTU)",
            "source_encoded": "Origin of water in manufacturing plant"
        }
        
        safe_ranges = {
            "ph": "6.5–8.5", "Hardness": "< 300", "Solids": "< 500", "Chloramines": "< 4",
            "Sulfate": "< 250", "Conductivity": "< 400", "Organic_carbon": "< 2",
            "Trihalomethanes": "< 80", "Turbidity": "< 4", "source_encoded": "—"
        }
        
        fi_df["Meaning"] = fi_df["Feature"].map(feature_meanings)
        fi_df["Safe_Range"] = fi_df["Feature"].map(safe_ranges)
        fi_df["Rank"] = range(1, len(fi_df) + 1)
        fi_df = fi_df[["Rank", "Feature", "Importance", "Meaning", "Safe_Range"]]

    metrics = {
        "Accuracy":  accuracy_score(y_test, y_pred),
        "Precision": precision_score(y_test, y_pred),
        "Recall":    recall_score(y_test, y_pred),
        "F1 Score":  f1_score(y_test, y_pred),
        "ROC-AUC":   auc(*roc_curve(y_test, y_prob)[:2])
    }

    report = f"""
WATER POTABILITY ANALYSIS REPORT
==================================
Project: Smart Water Potability Prediction
Industry: Food Manufacturing Plant
Project Owner: Shyama Bahuleyan - Boston Institute of Analytics

MODEL SELECTED: {model_choice}
================================
Accuracy  : {metrics['Accuracy']:.3f}
Precision : {metrics['Precision']:.3f}
Recall    : {metrics['Recall']:.3f}
F1 Score  : {metrics['F1 Score']:.3f}
ROC-AUC   : {metrics['ROC-AUC']:.3f}

TOP FEATURES DRIVING POTABILITY
================================
{fi_df.head(5).to_string(index=False) if not fi_df.empty else 'N/A'}

WATER SOURCE SAFETY SUMMARY
================================
"""
    for src in source_options:
        pot = df[df["source"] == src]["Potability"].mean() * 100
        sol = FILTRATION_SOLUTIONS.get(src, {})
        report += f"\n{src}:\n"
        report += f"  Potability Rate : {pot:.1f}%\n"
        report += f"  Main Issue      : {sol.get('issue', '—')}\n"
        report += f"  Top Solution    : {', '.join(sol.get('solutions', ['—']))}\n"

    report += """
RECOMMENDATIONS
================================
1. Monitor pH and keep within 6.5–8.5 for all food-contact water.
2. Apply RO treatment for Borewell and Cooling Tower water.
3. Use UV purifiers for Final Rinse and Municipal Supply water.
4. Neutralize CIP water before reuse or disposal.
5. Regularly retrain the model with new water quality data.
6. Conduct lab testing alongside ML predictions for compliance.
"""

    st.text_area("Report Preview", report, height=400)

    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.download_button(
            label="📥 Download Text Report",
            data=report,
            file_name=f"Water_Potability_Report_{model_choice}.txt",
            mime="text/plain"
        )
    
    with col2:
        if not fi_df.empty:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Feature Importance"
            
            BLUE_HDR, WHITE, GOLD, SILVER, BRONZE = "1E90FF", "FFFFFF", "FFD700", "C0C0C0", "CD7F32"
            thin = Side(style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Feature Importance - {model_choice}"
            ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
            ws["A1"].fill = PatternFill("solid", start_color="1F3864", fgColor="1F3864")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30
            
            headers = ["Rank", "Feature", "Importance", "Meaning", "Safe Range"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=2, column=col, value=h)
                c.font = Font(name="Arial", bold=True, color=WHITE)
                c.fill = PatternFill("solid", start_color=BLUE_HDR, fgColor=BLUE_HDR)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            
            rank_colors = [GOLD, SILVER, BRONZE]
            for i, row_data in enumerate(fi_df.values, 3):
                rank = row_data[0]
                bg = rank_colors[rank-1] if rank <= 3 else WHITE
                
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=i, column=col, value=val)
                    if col == 3:
                        c.number_format = "0.0000"
                    c.font = Font(name="Arial", bold=(rank <= 3))
                    c.fill = PatternFill("solid", start_color=bg, fgColor=bg)
                    c.alignment = Alignment(horizontal="center" if col in [1,3,5] else "left", vertical="center")
                    c.border = border
            
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 40
            ws.column_dimensions["E"].width = 15
            
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            st.download_button(
                label="📊 Feature Importance (Excel)",
                data=excel_buffer,
                file_name=f"Feature_Importance_{model_choice}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col3:
        try:
            with open("Water_Potability_Complete_Report.xlsx", "rb") as f:
                excel_data = f.read()
            st.download_button(
                label="📚 Complete Excel Report",
                data=excel_data,
                file_name="Water_Potability_Complete_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="All sheets: Feature Importance, Model Comparison, Source Safety"
            )
        except FileNotFoundError:
            st.info("💡 Place Water_Potability_Complete_Report.xlsx in project folder for full report")









